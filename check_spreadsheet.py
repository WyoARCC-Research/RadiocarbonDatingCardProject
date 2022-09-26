#check_spreadsheet.py contains functions used to verify and in some cases correct data entries in the excel file, it cannot verify 
#all fields
#this file also contains functions that were used to create a list of valid materials
#the checkallentries function is the most useful and will create seperate spreadsheets of all the entries it deems "good" and "bad"

#written by Collin Dixon

#This will NOT work on teton at the moment because it does not have openpyxl
from asyncio.windows_events import NULL
import re
import os
from tokenize import Double

#openpyxl should let us edit and access excel files
from openpyxl import Workbook
from openpyxl import load_workbook
#so we can mark cells by changing their font to red
from openpyxl.styles import Font
red = Font(color="FF0000")

#this way we can iterate over fields
fields={"Location: ":"F", "Material Dated: ":"C",
"Lab Name: ":"AF","Lab Number: ":"A","Age: ":"Q","Age Sigma: ":"R","Latitude: ":"G","Longitude: ":"H","Type Of Date: ":"E","Site Identifier: ":"X"}

#here is a place to store the materials when we use catalogMAterialDated
materials={}

#this function will check an entry for errors, checking fields from left to right. If an error is found it does not check the other fields
def checkEntry(row,worksheetIn):
    print(row)    
    #assume the entry is good until an error is found
    good_entry=True
    column=NULL
    #if statements for each implemented field check
    #when we find an error, return its column so we know where the error was
    if not checkLab_Number(row,worksheetIn):
        good_entry=False
        column='A'
        return [good_entry,column]
    
    if not checkMaterial(row,worksheetIn): 
        good_entry=False
        column='C'
        return [good_entry,column]

    if not checkTypeofDate(row, worksheetIn):
        good_entry=False
        column='E'
        return[good_entry,column]
    
    if not checklatlon(row,worksheetIn):
        good_entry=False
        column='G'
        return[good_entry,column]
    
    return [good_entry,column]



#this basically just calls checkentries interatively to check the whole document
#this is the main function to be using
def checkAllEntries(worksheet,start,end):
    
    #load new worksheet to store bad and good entries
    bad_wb=load_workbook('CARD Upload Template.xlsx')
    bad_ws =bad_wb['Data Fields']
    good_wb=load_workbook('CARD Upload Template.xlsx')
    good_ws =good_wb['Data Fields']
    
    #iterate over the rows in our range and check them, add them to good or ba appropriately
    #row counter to pass to checkentry
    rowcounter=start
    #finding the max_row of the worksheet is very slow so we do this instead. It is better to just provide a number to the script
    bad_output_current_row=5
    good_output_current_row=5
    
    for row in worksheet.iter_rows(min_row=start, max_col=33, max_row=end):
        #catalogMaterial_Dated(rowcounter,worksheet)
        output=checkEntry(rowcounter,worksheet)
        if not output[0]:#this is a bad entry!
            #if there is a bad row we copy all of its cells into our new spreadsheet       
            for cell in row:
                #for some reason it only works if we assign the address seperately like this
                addy=str(cell.column_letter)+str(bad_output_current_row)
                v=cell.value
                #assign the value to the corresponding cell
                bad_ws[addy]= v

            #change text to red so we can see what error was found
            #note lat and lon are checked together and will always turn lat red even if only lon is wrong
            bad_ws[str(output[1])+str(bad_output_current_row)].font=red
            bad_output_current_row=bad_output_current_row+1
        else: #the entry is good! (as far as the fields we can check are concerned)
            
            for cell in row:
                #for some reason it only works if we assign the address seperately like this
                addy=str(cell.column_letter)+str(good_output_current_row)
                v=cell.value
                #assign the value to the corresponding cell
                good_ws[addy]= v
            good_output_current_row=good_output_current_row+1
        rowcounter=rowcounter+1
            
    
    bad_wb.save("Bad_Cards.xlsx")
    good_wb.save("Good_Cards.xlsx")
    return

#will check the lab number of an entry
#this function may also attempt to correct the entry and change it
def checkLab_Number(row,worksheet):
    
    #lab numbers must be in AAA-nnn with varying numbers of characters according to the instructions of the Canadian radiocarbon database
    accepted_forms=["[A-Za-z]-[0-9][0-9][0-9][0-9]" #A-nnnn
                    ,"[A-Za-z]-[0-9][0-9][0-9]"#A-nnn
                    ,"[A-Za-z]-[0-9][0-9]"#A-nn
                    ,"[A-Za-z]-[0-9]"#A-n
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AA-nnnn
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AA-nnn
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9]"#AA-nn
                    ,"[A-Za-z][A-Za-z]-[0-9]"#AA-n
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AAA-nnnn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AAA-nnn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9]"#AAA-nn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9]"#AAA-n
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AAAA-nnnn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AAAA-nnn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9]"#AAAA-nn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9]"#AAAA-n
                    ]
    text=worksheet['A'+str(row)].value
    
    #use found to track if text is in accepted form
    
    found=False
    #cycle through accepted forms, if it is in an accepted form then we are done. if not, then found=False and we move on to check for any matches in the text
    # , not just exact matches
    for regex in accepted_forms:
        try:
            match=re.fullmatch(regex,text)
        except:
            break
        if match !=None:
            found =re.fullmatch(regex,text).group(0)
            break

    #if we didnt find an exact match, we start looking for parts of the sring that match, and take the longest of these as the match
    
    if not found:
        potential_match=""
        for regex in accepted_forms:        
            try:
                search=re.search(regex,text)

                
                if len(search.group(0))>len(potential_match):
                    potential_match=search.group(0)
            except:
                pass

        if len(potential_match)>0:
            worksheet['A'+str(row)].value=potential_match
            found=potential_match
            #keep track of what we extracted from its larger string
            extractedLabNumbers.write("Card "+worksheet['AC'+str(row)].value+" : "+found+ " out of "+text+os.linesep)  
    
    #then if that didnt work, remove spaces and look for matches
    if not found:
        try:
            newtext=text.replace(" ","")
            potential_match=""
            for regex in accepted_forms:        
                try:
                    search=re.search(regex,newtext)

                    
                    if len(search.group(0))>len(potential_match):
                        potential_match=search.group(0)
                except:
                    pass

            if len(potential_match)>0:
                worksheet['A'+str(row)].value=potential_match
                found=potential_match
                #keep track of what we extracted from its larger string
                extractedLabNumbers.write("Card "+worksheet['AC'+str(row)].value+" : "+found+ " out of "+text+os.linesep) 
        except:
            pass 
    
    return found

#references a material against the valid materials to check its validity
def checkMaterial(row, worksheet):
    text=worksheet['C'+str(row)].value
    
    try:
        text=text.lower()
        text=text.replace(" ","")
        text=text.replace("-","")
        text=text.replace("_","")
        text=text.replace("+","")
        text=text.replace("^","")
        text=text.replace("•","")
        text=text.replace(",","")
        text=text.replace("[","")
        text=text.replace("]","")
        text=text.replace("(","")
        text=text.replace(")","")
        text=text.replace(".","")
    except:
        return False
    found=False

    try:
        if validMaterials[text]:
            found=text
    except:
        found=False

    return found

#there are only 3 valid date types,  "archaeological", paleontological", and "geological"
def checkTypeofDate(row, worksheet):
    text=worksheet['E'+str(row)].value
    try:
        text=text.lower()
    except:
        return False
    
    validText=["archaeology", "paleontology","geology"]
    for entry in validText:
        if text==entry:
            return True

    return False

#note that since these are being checked together, only an incorrect lat will be turned red in the spreadsheet
def checklatlon(row, worksheet):
    text1=worksheet['G'+str(row)].value
    text2=worksheet['H'+str(row)].value
    try:
        text1=float(text1)       
        text2=float(text2)
    except:
        return False

    if(text1>90 or text1<-90 or text2>180 or text2<-180):
        return False
    else:
        return True


#--------------------------------------------------------------------------

#these material functions are mostly useless now since we have already created a valid materials list
#however they may be useful if a new list of materials ever needs to be created.

#this function takes a row, finds its material and catalogs in in a text document
#this way we can see all of the materials that are being read in
def catalogMaterial_Dated(row,worksheetIn):
    
    text=worksheetIn['C'+str(row)].value

    #in order for case insensitive comparison, we also need to make sure that it actually has a material field and if not, record that
    try:
        lowerStr=text.lower()
        lowerStr=lowerStr.replace(" ","")
        lowerStr=lowerStr.replace("-","")
        lowerStr=lowerStr.replace("_","")
        lowerStr=lowerStr.replace("+","")
        lowerStr=lowerStr.replace("^","")
        lowerStr=lowerStr.replace("•","")
        lowerStr=lowerStr.replace(",","")
        lowerStr=lowerStr.replace("[","")
        lowerStr=lowerStr.replace("]","")
        lowerStr=lowerStr.replace("(","")
        lowerStr=lowerStr.replace(")","")
        lowerStr=lowerStr.replace(".","")
        
    #some entries will have strange characters so we catch them here
    except:
        lowerStr='materialfielderror'
        

    #these are the materials we have found so far
    mats=materials.keys()
    #iterate through keys and compare without regard for case, keep a tally of each material
    found=False
    for key in mats:
        if key==lowerStr:
            materials[key]=materials[key]+1
            found=True
            break
    if not found:
        materials[lowerStr]=1

    return

#sorts and lists the materials in the materials dict to a text file
def listMaterials():

    file = open("DictFile.txt","w",encoding="utf-8")
    
    for w in sorted(materials, key=materials.get, reverse=True):
        file.write('%s:%s\n' % (w, materials[w]))
    
    file.close()
    return

#reads in a dict of valid materials from a text file
def getValidMaterials(filename):
    materialsFile=open(filename,encoding="utf-8")
    validMaterials={}
    
    #we get rid of the tally
    for line in materialsFile.readlines():
        text=line.split(':')[0]
        validMaterials[text]=True
    return validMaterials






#this will use the materials and crossrefererences with the valid materials to find what I manually removed from validmaterials
def generateRemoved():
    removedMats = open("Removed_materials.txt","w",encoding="utf-8")
    for key in materials.keys():
        try:
            x=validMaterials[key]
        except:
            removedMats.write(key+os.linesep)
    removedMats.close()



wb=load_workbook('Card_Upload_Template_Output2.xlsx')
ws =wb['Data Fields']
extractedLabNumbers = open("Extracted_Lab_Numbers.txt","w",encoding="utf-8")

validMaterials=getValidMaterials("valid_materials.txt")




checkAllEntries(ws,5,ws.max_row)


extractedLabNumbers.close()



print('done')

